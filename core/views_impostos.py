from django.shortcuts import render, redirect
from django.contrib import messages
from .models import TaxaImposto
from decimal import Decimal
import openpyxl
from django.http import HttpResponse

def taxas_imposto_list(request):
    """Visualiza a tabela de impostos e taxas em vigor"""
    taxas = TaxaImposto.objects.all().order_by('nome')
    return render(request, 'contabilidade/taxas_imposto.html', {'taxas': taxas})

def atualizar_taxas_padrao(request):
    """Reseta a Tabela de Impostos para os valores padrão de Angola"""
    if request.method == 'POST':
        # Mantém ou atualiza os existentes baseados no código
        dados = [
            ('IVA_NORMAL', 'IVA - Taxa Geral', Decimal('14.0'), 'Taxa normal aplicada à maioria dos bens e serviços.'),
            ('IVA_SIMP', 'IVA - Regime Simplificado', Decimal('7.0'), 'Aplicado a empresas no regime simplificado.'),
            ('IVA_RESTAURACAO', 'IVA - Hotelaria e Restauração', Decimal('7.0'), 'Taxa reduzida para o sector de hotelaria.'),
            ('IND_NORMAL', 'Imposto Industrial - Geral', Decimal('25.0'), 'Taxa padrão aplicada aos lucros das empresas.'),
            ('IND_FINANC', 'Imposto Industrial - Bancos/Seguros', Decimal('35.0'), 'Taxa para instituições financeiras.'),
            ('IND_AGRO', 'Imposto Industrial - Agricultura/Pecuária', Decimal('10.0'), 'Taxa reduzida para incentivo ao sector primário.'),
            ('RF_SERVICOS', 'Retenção na Fonte - Serviços (6.5%)', Decimal('6.5'), 'Retenção aplicada no pagamento de serviços a residentes.'),
            ('RF_IPU_RENDAS', 'Retenção na Fonte - IPU Rendas (15%)', Decimal('15.0'), 'Retenção de IPU sobre o valor das rendas de imóveis.'),
            ('SELO_RECIBO', 'Imposto de Selo - Recibos/Quitação', Decimal('1.0'), 'Aplicado sobre o valor total do recibo.'),
            ('INSS_FUNC', 'INSS - Parcela do Trabalhador', Decimal('3.0'), 'Desconto directo no salário do funcionário.'),
            ('INSS_EMP', 'INSS - Parcela da Empresa', Decimal('8.0'), 'Contribuição a cargo da entidade empregadora.'),
        ]
        
        for codigo, nome, taxa, desc in dados:
            TaxaImposto.objects.update_or_create(
                codigo=codigo,
                defaults={'nome': nome, 'taxa': taxa, 'descricao': desc, 'ativo': True}
            )
            
        messages.success(request, 'Taxas de imposto atualizadas com sucesso para o padrão de Angola.')
    
    return redirect('taxas_imposto_list')

def exportar_taxas_excel(request):
    """Exporta a tabela de taxas para um ficheiro Excel"""
    taxas = TaxaImposto.objects.all().order_by('nome')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Taxas de Imposto"
    
    # Cabeçalho da Empresa
    empresa = request.empresa
    ws['A1'] = empresa.nome
    ws['A1'].font = openpyxl.styles.Font(bold=True, size=12)
    ws['A2'] = f"NIF: {empresa.nif or '---'}"
    ws['A3'] = f"Data de Emissão: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    
    ws['A5'] = "TABELA DE TAXAS DE IMPOSTO"
    ws['A5'].font = openpyxl.styles.Font(bold=True, size=10)
    
    # Cabeçalhos da Tabela
    headers = ["Nome do Imposto", "Código", "Taxa (%)", "Descrição", "Estado", "Última Atualização"]
    ws.append([]) # Linha vazia
    ws.append(headers)
    
    # Estilo básico para cabeçalhos da tabela
    header_row = 7
    for cell in ws[header_row]:
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Dados
    for t in taxas:
        ws.append([
            t.nome,
            t.codigo,
            float(t.taxa),
            t.descricao or "",
            "Em Vigor" if t.ativo else "Suspenso",
            t.ultima_atualizacao.strftime("%d/%m/%Y %H:%M")
        ])
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=taxas_imposto_angola.xlsx'
    wb.save(response)
    return response
