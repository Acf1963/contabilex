

from django.http import HttpResponse
from io import BytesIO
from datetime import datetime
from .models import Cliente, Fornecedor


# ===== EXPORTAÇÃO DE CLIENTES E FORNECEDORES =====

def exportar_clientes_excel(request):
    """Exportar lista de clientes para Excel"""
    empresa = request.empresa
    clientes = Cliente.objects.filter(empresa=empresa).order_by('codigo_contabilistico')
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return exportar_clientes_csv(request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Headers
    headers = ["Código PGC", "Nome", "NIF", "Email", "Telefone", "Endereço"]
    ws.append(headers)
    
    # Estilo do header
    header_fill = PatternFill(start_color="58A6FF", end_color="58A6FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dados
    for cliente in clientes:
        ws.append([
            cliente.codigo_contabilistico,
            cliente.nome,
            cliente.nif or "",
            cliente.email or "",
            cliente.telefone or "",
            cliente.endereco or "",
        ])
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    
    # Retornar arquivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response

def exportar_clientes_csv(request):
    """Exportar lista de clientes para CSV (fallback)"""
    import csv
    empresa = request.empresa
    clientes = Cliente.objects.filter(empresa=empresa).order_by('codigo_contabilistico')
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="Clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(["Código PGC", "Nome", "NIF", "Email", "Telefone", "Endereço"])
    
    for cliente in clientes:
        writer.writerow([
            cliente.codigo_contabilistico,
            cliente.nome,
            cliente.nif or "",
            cliente.email or "",
            cliente.telefone or "",
            cliente.endereco or "",
        ])
    
    return response

def exportar_clientes_pdf(request):
    """Exportar lista de clientes para PDF"""
    empresa = request.empresa
    clientes = Cliente.objects.filter(empresa=empresa).order_by('codigo_contabilistico')
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
    except ImportError:
        return HttpResponse("ReportLab não está instalado. Instale com: pip install reportlab", status=400)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#58A6FF'),
        spaceAfter=12,
        alignment=1  # Center
    )
    elements.append(Paragraph(f"Lista de Clientes - {empresa.nome}", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Tabela
    data = [["Código PGC", "Nome", "NIF", "Telefone"]]
    for cliente in clientes:
        data.append([
            cliente.codigo_contabilistico,
            cliente.nome[:30],  # Limitar comprimento
            cliente.nif or "",
            cliente.telefone or "",
        ])
    
    table = Table(data, colWidths=[1.2*inch, 2*inch, 1.2*inch, 1.3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#58A6FF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response

def exportar_fornecedores_excel(request):
    """Exportar lista de fornecedores para Excel"""
    empresa = request.empresa
    fornecedores = Fornecedor.objects.filter(empresa=empresa).order_by('codigo_contabilistico')
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return exportar_fornecedores_csv(request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fornecedores"
    
    # Headers
    headers = ["Código PGC", "Nome", "NIF", "Email", "Telefone", "Endereço"]
    ws.append(headers)
    
    # Estilo do header
    header_fill = PatternFill(start_color="34A853", end_color="34A853", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dados
    for fornecedor in fornecedores:
        ws.append([
            fornecedor.codigo_contabilistico,
            fornecedor.nome,
            fornecedor.nif or "",
            fornecedor.email or "",
            fornecedor.telefone or "",
            fornecedor.endereco or "",
        ])
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    
    # Retornar arquivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Fornecedores_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response

def exportar_fornecedores_csv(request):
    """Exportar lista de fornecedores para CSV (fallback)"""
    import csv
    empresa = request.empresa
    fornecedores = Fornecedor.objects.filter(empresa=empresa).order_by('codigo_contabilistico')
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="Fornecedores_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(["Código PGC", "Nome", "NIF", "Email", "Telefone", "Endereço"])
    
    for fornecedor in fornecedores:
        writer.writerow([
            fornecedor.codigo_contabilistico,
            fornecedor.nome,
            fornecedor.nif or "",
            fornecedor.email or "",
            fornecedor.telefone or "",
            fornecedor.endereco or "",
        ])
    
    return response

def exportar_fornecedores_pdf(request):
    """Exportar lista de fornecedores para PDF"""
    empresa = request.empresa
    fornecedores = Fornecedor.objects.filter(empresa=empresa).order_by('codigo_contabilistico')
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
    except ImportError:
        return HttpResponse("ReportLab não está instalado. Instale com: pip install reportlab", status=400)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#34A853'),
        spaceAfter=12,
        alignment=1  # Center
    )
    elements.append(Paragraph(f"Lista de Fornecedores - {empresa.nome}", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Tabela
    data = [["Código PGC", "Nome", "NIF", "Telefone"]]
    for fornecedor in fornecedores:
        data.append([
            fornecedor.codigo_contabilistico,
            fornecedor.nome[:30],  # Limitar comprimento
            fornecedor.nif or "",
            fornecedor.telefone or "",
        ])
    
    table = Table(data, colWidths=[1.2*inch, 2*inch, 1.2*inch, 1.3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34A853')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Fornecedores_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response
