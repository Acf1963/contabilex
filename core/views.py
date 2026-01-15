from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.db.models import Q
from .models import Cliente, Fornecedor, Fatura, Compra, Despesa, Empresa, LancamentoDiario, Conta
from .forms import (
    ClienteForm, FornecedorForm, FaturaForm, ItemFaturaFormSet, 
    CompraForm, ItemCompraFormSet, DespesaForm, EmpresaForm,
    LancamentoAberturaForm, MovimentoAberturaFormSet, UsuarioForm
)
from datetime import datetime
from io import BytesIO
import csv

import requests
import re
from django.core.cache import cache

def _get_bna_rates():
    """Busca taxas de câmbio através da API oficial do BNA"""
    rates = cache.get('bna_exchange_rates')
    if rates:
        return rates

    # Valores padrão de fallback
    usd = "912.286"
    eur = "1069.522"

    try:
        headers = {'User-Agent': 'Mozilla/5.0'}
        # API do BNA para taxas de referência (Venda Tipo B)
        url_base = "https://www.bna.ao/service/rest/taxas/get/taxa/referencia?tipocambio=b&moeda="
        
        import urllib3
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
        
        # USD
        res_usd = requests.get(f"{url_base}usd", headers=headers, timeout=8, verify=False)
        if res_usd.status_code == 200:
            data = res_usd.json()
            if data.get('success') and data.get('genericResponse'):
                usd = str(data['genericResponse'][0]['taxa'])
        
        # EUR
        res_eur = requests.get(f"{url_base}eur", headers=headers, timeout=8, verify=False)
        if res_eur.status_code == 200:
            data = res_eur.json()
            if data.get('success') and data.get('genericResponse'):
                eur = str(data['genericResponse'][0]['taxa'])
                
        rates = {'USD': usd, 'EUR': eur}
        cache.set('bna_exchange_rates', rates, 3600) # Cache por 1 hora
        return rates
    except Exception:
        # Silencioso no dashboard para não travar o carregamento
        return {'USD': usd, 'EUR': eur}

def home(request):
    """Dashboard principal com dados reais e saldos contabilísticos"""
    from django.db.models import Sum, Q
    from .models import MovimentoRazao, Conta
    
    empresa = request.empresa
    if not empresa:
        return redirect('configuracao_empresa')

    # 1. Saldos Disponíveis (PGC Classe 4)
    # Caixa (45)
    saldo_caixa = MovimentoRazao.objects.filter(
        lancamento__empresa=empresa,
        conta__codigo__startswith='45'
    ).aggregate(
        total=Sum('valor', filter=Q(tipo='D')) - Sum('valor', filter=Q(tipo='C'))
    )['total'] or 0
    
    # Bancos (43)
    saldo_bancos = MovimentoRazao.objects.filter(
        lancamento__empresa=empresa,
        conta__codigo__startswith='43'
    ).aggregate(
        total=Sum('valor', filter=Q(tipo='D')) - Sum('valor', filter=Q(tipo='C'))
    )['total'] or 0
    
    # 2. Saldos de Terceiros (PGC Classe 3)
    # Clientes (31) - Natureza Devedora
    saldo_clientes = MovimentoRazao.objects.filter(
        lancamento__empresa=empresa,
        conta__codigo__startswith='31'
    ).aggregate(
        total=Sum('valor', filter=Q(tipo='D')) - Sum('valor', filter=Q(tipo='C'))
    )['total'] or 0
    
    # Fornecedores (32) - Natureza Credora
    saldo_fornecedores = MovimentoRazao.objects.filter(
        lancamento__empresa=empresa,
        conta__codigo__startswith='32'
    ).aggregate(
        total=Sum('valor', filter=Q(tipo='C')) - Sum('valor', filter=Q(tipo='D'))
    )['total'] or 0
    
    # 3. Proveitos e Custos (PGC Classe 7 e 6) - Ano Corrente
    proveitos_total = MovimentoRazao.objects.filter(
        lancamento__empresa=empresa,
        conta__codigo__startswith='7'
    ).aggregate(
        total=Sum('valor', filter=Q(tipo='C')) - Sum('valor', filter=Q(tipo='D'))
    )['total'] or 0
    
    custos_total = MovimentoRazao.objects.filter(
        lancamento__empresa=empresa,
        conta__codigo__startswith='6'
    ).aggregate(
        total=Sum('valor', filter=Q(tipo='D')) - Sum('valor', filter=Q(tipo='C'))
    )['total'] or 0
    
    resultado_estimado = proveitos_total - custos_total
    margem_bruta = (resultado_estimado / proveitos_total * 100) if proveitos_total > 0 else 0
    
    # Dados para listagens
    ultimas_faturas = Fatura.objects.filter(empresa=empresa).order_by('-data_emissao')[:3]
    ultimas_despesas = Despesa.objects.filter(empresa=empresa).order_by('-data')[:3]
    total_clientes = Cliente.objects.filter(empresa=empresa).count()
    # 6. Taxas BNA
    bna_rates = _get_bna_rates()
    
    context = {
        'saldo_caixa': saldo_caixa,
        'saldo_bancos': saldo_bancos,
        'saldo_clientes': saldo_clientes,
        'saldo_fornecedores': saldo_fornecedores,
        'proveitos_total': proveitos_total,
        'custos_total': custos_total,
        'resultado_estimado': resultado_estimado,
        'margem_bruta': margem_bruta,
        'total_clientes': total_clientes,
        'ultimas_faturas': ultimas_faturas,
        'ultimas_despesas': ultimas_despesas,
        'bna_rates': bna_rates,
    }
    
    return render(request, 'dashboard.html', context)

def clientes_list(request):
    """Lista de Clientes e Formulário de Adição"""
    empresa = request.empresa
    clientes = Cliente.objects.filter(empresa=empresa).order_by('-id')
    
    if request.method == 'POST':
        form = ClienteForm(request.POST, empresa=empresa)
        if form.is_valid():
            cliente = form.save(commit=False)
            cliente.empresa = empresa
            cliente.save()
            messages.success(request, f'Cliente {cliente.nome} adicionado com o código {cliente.codigo_contabilistico}!')
            return redirect('clientes')
        else:
            messages.error(request, 'Erro ao adicionar cliente. Verifique os campos.')
    else:
        form = ClienteForm(empresa=empresa)

    return render(request, 'clientes.html', {
        'clientes': clientes,
        'form': form
    })

def editar_cliente(request, pk):
    """Edição de um cliente existente"""
    cliente = Cliente.objects.get(pk=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente, empresa=request.empresa)
        if form.is_valid():
            form.save()
            messages.success(request, f'Cliente {cliente.nome} atualizado com sucesso!')
            return redirect('clientes')
    else:
        form = ClienteForm(instance=cliente, empresa=request.empresa)
    
    return render(request, 'clientes_form.html', {
        'form': form,
        'cliente': cliente,
        'titulo': f'Editar Cliente: {cliente.nome}'
    })

def excluir_cliente(request, pk):
    """Excluir um cliente"""
    cliente = Cliente.objects.get(pk=pk)
    nome = cliente.nome
    try:
        cliente.delete()
        messages.success(request, f'Cliente {nome} removido com sucesso.')
    except Exception as e:
        messages.error(request, f'Não foi possível remover o cliente. Ele pode estar associado a faturas.')
    
    return redirect('clientes')

def fornecedores_list(request):
    """Lista de Fornecedores e Formulário de Adição"""
    empresa = request.empresa
    fornecedores = Fornecedor.objects.filter(empresa=empresa).order_by('-id')
    
    if request.method == 'POST':
        form = FornecedorForm(request.POST, empresa=empresa)
        if form.is_valid():
            fornecedor = form.save(commit=False)
            fornecedor.empresa = empresa
            fornecedor.save()
            messages.success(request, f'Fornecedor {fornecedor.nome} adicionado com o código {fornecedor.codigo_contabilistico}!')
            return redirect('fornecedores')
        else:
            messages.error(request, 'Erro ao adicionar fornecedor. Verifique os campos.')
    else:
        form = FornecedorForm(empresa=empresa)

    return render(request, 'fornecedores.html', {
        'fornecedores': fornecedores,
        'form': form
    })

def editar_fornecedor(request, pk):
    """Edição de um fornecedor existente"""
    fornecedor = Fornecedor.objects.get(pk=pk)
    if request.method == 'POST':
        form = FornecedorForm(request.POST, instance=fornecedor, empresa=request.empresa)
        if form.is_valid():
            form.save()
            messages.success(request, f'Fornecedor {fornecedor.nome} atualizado com sucesso!')
            return redirect('fornecedores')
    else:
        form = FornecedorForm(instance=fornecedor, empresa=request.empresa)
    
    return render(request, 'fornecedores_form.html', {
        'form': form,
        'fornecedor': fornecedor,
        'titulo': f'Editar Fornecedor: {fornecedor.nome}'
    })

def excluir_fornecedor(request, pk):
    """Excluir um fornecedor"""
    fornecedor = Fornecedor.objects.get(pk=pk)
    nome = fornecedor.nome
    try:
        fornecedor.delete()
        messages.success(request, f'Fornecedor {nome} removido com sucesso.')
    except Exception as e:
        messages.error(request, f'Não foi possível remover o fornecedor. Ele pode estar associado a despesas.')
    
    return redirect('fornecedores')

def faturas_list(request):
    empresa = request.empresa
    faturas = Fatura.objects.filter(empresa=empresa).order_by('-id')
    return render(request, 'faturas_lista.html', {'faturas': faturas})

def nova_fatura(request):
    empresa = request.empresa
    if request.method == 'POST':
        form = FaturaForm(request.POST, empresa=empresa)
        formset = ItemFaturaFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            fatura = form.save(commit=False)
            fatura.empresa = empresa
            fatura.save() # Generate Number
            
            items = formset.save(commit=False)
            for item in items:
                item.fatura = fatura
                item.save()
            
            fatura.calcular_totais()
            messages.success(request, f'Fatura {fatura.numero} emitida com sucesso!')
            return redirect('faturas_list')
    else:
        form = FaturaForm(empresa=empresa)
        formset = ItemFaturaFormSet()

    return render(request, 'fatura_nova.html', {
        'form': form,
        'formset': formset
    })

def fatura_detalhe(request, pk):
    """Visualiza detalhes da fatura e seus lançamentos contabilisticos"""
    from .forms import PagamentoFaturaForm
    fatura = Fatura.objects.prefetch_related('itens', 'lancamentos__movimentos__conta', 'pagamentos').get(pk=pk)
    form_pagamento = PagamentoFaturaForm(initial={'valor': fatura.saldo_pendente})
    return render(request, 'fatura_detalhe.html', {
        'fatura': fatura,
        'form_pagamento': form_pagamento
    })

def registrar_pagamento_fatura(request, pk):
    """Regista um pagamento (total ou parcial) para uma fatura"""
    from django.shortcuts import get_object_or_404
    from .forms import PagamentoFaturaForm
    from .models import LancamentoDiario, MovimentoRazao, Conta
    
    fatura = get_object_or_404(Fatura, pk=pk)
    if request.method == 'POST':
        form = PagamentoFaturaForm(request.POST)
        if form.is_valid():
            pagamento = form.save(commit=False)
            pagamento.fatura = fatura
            pagamento.save()
            
            # Atualizar estado da fatura
            if fatura.saldo_pendente <= 0:
                fatura.estado = 'PAGA'
            else:
                fatura.estado = 'PAGAMENTO_PARCIAL'
            fatura.save()
            
            # Lançamento Contabilístico do Pagamento
            # Débito: 43 (Banco) ou 45 (Caixa)
            # Crédito: 31 (Cliente - conta específica)
            if request.POST.get('gerar_lancamento'):
                try:
                    codigo_conta_recebimento = '43' if pagamento.metodo_pagamento == 'BANCO' else '45'
                    # Buscar conta de recebimento (global ou empresa)
                    conta_recebimento = Conta.objects.filter(codigo=codigo_conta_recebimento).filter(Q(empresa=fatura.empresa) | Q(empresa__isnull=True)).first()
                    
                    # Buscar conta do cliente (específica da empresa)
                    conta_cliente = Conta.objects.filter(codigo=fatura.cliente.codigo_contabilistico, empresa=fatura.empresa).first()
                    
                    if conta_recebimento and conta_cliente:
                        lancamento = LancamentoDiario.objects.create(
                            empresa=fatura.empresa,
                            data=pagamento.data,
                            descricao=f"Recebimento {fatura.numero} - Parcela",
                            fatura=fatura
                        )
                    
                        # Débito (Dinheiro entra)
                        MovimentoRazao.objects.create(
                            lancamento=lancamento,
                            conta=conta_recebimento,
                            tipo='D',
                            valor=pagamento.valor
                        )
                        
                        # Crédito (Dívida do cliente diminui)
                        MovimentoRazao.objects.create(
                            lancamento=lancamento,
                            conta=conta_cliente,
                            tipo='C',
                            valor=pagamento.valor
                        )
                except Exception as e:
                    messages.warning(request, f'Pagamento registado, mas erro ao gerar lançamento contabilístico: {str(e)}')
            
            messages.success(request, f'Pagamento de Kz {pagamento.valor} registado com sucesso!')
            
    return redirect('fatura_detalhe', pk=pk)

def mudar_estado_fatura(request, pk, estado):
    """Muda o estado de uma fatura e desencadeia lançamentos se necessário"""
    from .models import Conta, LancamentoDiario, MovimentoRazao, PagamentoFatura
    fatura = Fatura.objects.get(pk=pk)
    antigo_estado = fatura.estado
    
    if estado in dict(Fatura.ESTADOS):
        # Se mudar para PAGA manualmente (e não estava PAGA), registar pagamento total
        if estado == 'PAGA' and antigo_estado != 'PAGA':
            saldo = fatura.saldo_pendente
            if saldo > 0:
                PagamentoFatura.objects.create(
                    fatura=fatura,
                    valor=saldo,
                    metodo_pagamento='BANCO',
                    observacoes="Pagamento total (alteração manual de estado)"
                )
                # O lançamento do pagamento será gerado pela lógica de sinal ou manual? 
                # Por agora, vamos simplificar e focar no estado.
        
        fatura.estado = estado
        fatura.save()
        
        # Lógica de Emissão (Lançamento Diário)
        if estado == 'EMITIDA' and antigo_estado == 'RASCUNHO':
            try:
                # Buscar conta do cliente
                conta_cliente = Conta.objects.filter(
                    codigo=fatura.cliente.codigo_contabilistico,
                    empresa=fatura.empresa
                ).first()
                
                # Procura conta de proveitos (Classe 7)
                conta_proveitos = Conta.objects.filter(
                    codigo__startswith='7',
                    aceita_lancamentos=True
                ).filter(Q(empresa=fatura.empresa) | Q(empresa__isnull=True)).first()
                
                if conta_cliente and conta_proveitos:
                    lancamento = LancamentoDiario.objects.create(
                        empresa=fatura.empresa,
                        data=fatura.data_emissao,
                        descricao=f"Fatura {fatura.numero} - Emissão",
                        fatura=fatura
                    )
                    
                    # 1. Débito no Cliente (Valor a Pagar = Total - Retenção)
                    MovimentoRazao.objects.create(lancamento=lancamento, conta=conta_cliente, tipo='D', valor=fatura.total_a_pagar)
                    
                    # 2. Se houver retenção, Debitar conta de Retenção a Recuperar pelo Cliente ou Estado (Ativo)
                    # O user pediu para criar "conta para o cliente com retenção". 
                    # Vamos assumir uma conta 31.8 "Clientes - Retenção" ou similar, se não existir fallback para 34.
                    if fatura.valor_retencao > 0:
                        conta_retencao = Conta.objects.filter(codigo='31.8', empresa=fatura.empresa).first() 
                        if not conta_retencao:
                            conta_retencao = Conta.objects.filter(codigo__startswith='34', empresa=fatura.empresa).first() # Fallback genérico Estado
                        
                        if conta_retencao:
                            MovimentoRazao.objects.create(lancamento=lancamento, conta=conta_retencao, tipo='D', valor=fatura.valor_retencao)
                    
                    # 3. Crédito em Proveitos/Vendas (Valor base ou total? PGC = Total Proveito)
                    # Na prática contabilistica PGC: Vendas (Credito) líquidas de imposto se houver IVA.
                    # Aqui stamos a lançar o Total Bruto em vendas (simplificado se não configurado IVA separado).
                    # Se tiver Imposto separado:
                    if fatura.total_imposto > 0:
                        conta_iva = Conta.objects.filter(codigo='34.5.3', empresa=fatura.empresa).first() # IVA Liquidado
                        if not conta_iva: conta_iva = Conta.objects.filter(codigo__startswith='34', empresa=fatura.empresa).first()
                        
                        MovimentoRazao.objects.create(lancamento=lancamento, conta=conta_proveitos, tipo='C', valor=fatura.subtotal)
                        if conta_iva:
                            MovimentoRazao.objects.create(lancamento=lancamento, conta=conta_iva, tipo='C', valor=fatura.total_imposto)
                    else:
                        MovimentoRazao.objects.create(lancamento=lancamento, conta=conta_proveitos, tipo='C', valor=fatura.total)
            except Exception as e:
                pass # Silencioso se não houver contas configuradas
                
        messages.success(request, f'Estado da fatura {fatura.numero} alterado para {estado}.')
    return redirect('faturas_list')

def compras_list(request):
    empresa = request.empresa
    compras = Compra.objects.filter(empresa=empresa).order_by('-id')
    return render(request, 'compras_lista.html', {'compras': compras})

def nova_compra(request):
    empresa = request.empresa
    if request.method == 'POST':
        form = CompraForm(request.POST, empresa=empresa)
        formset = ItemCompraFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            compra = form.save(commit=False)
            compra.empresa = empresa
            compra.save() # Generate Number
            
            items = formset.save(commit=False)
            for item in items:
                item.compra = compra
                item.save()
            
            compra.calcular_totais()
            messages.success(request, f'Compra {compra.numero} registada com sucesso!')
            return redirect('compras_list')
    else:
        form = CompraForm(empresa=empresa)
        formset = ItemCompraFormSet()

    return render(request, 'compra_nova.html', {
        'form': form,
        'formset': formset
    })

def compra_detalhe(request, pk):
    """Visualiza detalhes da compra e seus lançamentos contabilisticos"""
    from .forms import PagamentoCompraForm
    compra = Compra.objects.prefetch_related('itens', 'lancamentos__movimentos__conta', 'pagamentos').get(pk=pk)
    form_pagamento = PagamentoCompraForm(initial={'valor': compra.saldo_pendente})
    return render(request, 'compra_detalhe.html', {
        'compra': compra,
        'form_pagamento': form_pagamento
    })

def registrar_pagamento_compra(request, pk):
    """Regista um pagamento (total ou parcial) para uma compra"""
    from django.shortcuts import get_object_or_404, redirect
    from .forms import PagamentoCompraForm
    from .models import Compra, LancamentoDiario, MovimentoRazao, Conta
    
    compra = get_object_or_404(Compra, pk=pk)
    if request.method == 'POST':
        form = PagamentoCompraForm(request.POST)
        if form.is_valid():
            pagamento = form.save(commit=False)
            pagamento.compra = compra
            pagamento.save()
            
            # Atualizar estado da compra
            if compra.saldo_pendente <= 0:
                compra.estado = 'PAGA'
            else:
                compra.estado = 'PAGAMENTO_PARCIAL'
            compra.save()
            
            # Lançamento Contabilístico do Pagamento
            # Débito: 32 (Fornecedor - conta específica) - Reduz dívida
            # Crédito: 43 (Banco) ou 45 (Caixa) - Sai dinheiro
            if request.POST.get('gerar_lancamento'):
                try:
                    codigo_conta_pagamento = '43' if pagamento.metodo_pagamento == 'BANCO' else '45'
                    # Buscar conta de pagamento (global ou empresa)
                    conta_pagamento = Conta.objects.filter(codigo=codigo_conta_pagamento).filter(Q(empresa=compra.empresa) | Q(empresa__isnull=True)).first()
                    
                    # Buscar conta do fornecedor (específica da empresa)
                    conta_fornecedor = Conta.objects.filter(codigo=compra.fornecedor.codigo_contabilistico, empresa=compra.empresa).first()
                    
                    if conta_pagamento and conta_fornecedor:
                        lancamento = LancamentoDiario.objects.create(
                            empresa=compra.empresa,
                            data=pagamento.data,
                            descricao=f"Pagamento {compra.numero} - Parcela",
                            compra=compra
                        )
                    
                        # Débito (Dívida do fornecedor diminui)
                        MovimentoRazao.objects.create(
                            lancamento=lancamento,
                            conta=conta_fornecedor,
                            tipo='D',
                            valor=pagamento.valor
                        )
                        
                        # Crédito (Dinheiro sai)
                        MovimentoRazao.objects.create(
                            lancamento=lancamento,
                            conta=conta_pagamento,
                            tipo='C',
                            valor=pagamento.valor
                        )
                except Exception as e:
                    messages.warning(request, f'Pagamento registado, mas erro ao gerar lançamento contabilístico: {str(e)}')
            
            messages.success(request, f'Pagamento de Kz {pagamento.valor} registado com sucesso!')
            
    return redirect('compra_detalhe', pk=pk)

def mudar_estado_compra(request, pk, estado):
    """Muda o estado de uma compra e gera lançamentos se necessário"""
    from .models import Conta, LancamentoDiario, MovimentoRazao, Compra, PagamentoCompra
    compra = Compra.objects.get(pk=pk)
    antigo_estado = compra.estado
    
    if estado in dict(Compra.ESTADOS):
        # Se mudar para PAGA manualmente (e não estava PAGA), registar pagamento total
        if estado == 'PAGA' and antigo_estado != 'PAGA':
            saldo = compra.saldo_pendente
            if saldo > 0:
                PagamentoCompra.objects.create(
                    compra=compra,
                    valor=saldo,
                    metodo_pagamento='BANCO',
                    observacoes="Pagamento total (alteração manual de estado)"
                )
        
        compra.estado = estado
        compra.save()
        
        # Lógica de Emissão/Registo (Lançamento Diário)
        if estado == 'REGISTADA' and antigo_estado == 'RASCUNHO':
            try:
                # Buscar conta do fornecedor
                conta_fornecedor = Conta.objects.filter(
                    codigo=compra.fornecedor.codigo_contabilistico,
                    empresa=compra.empresa
                ).first()
                
                # Procura conta de custos/gastos (Classe 6)
                conta_gastos = Conta.objects.filter(
                    codigo__startswith='6',
                    aceita_lancamentos=True
                ).filter(Q(empresa=compra.empresa) | Q(empresa__isnull=True)).first()
                
                if conta_fornecedor and conta_gastos:
                    lancamento = LancamentoDiario.objects.create(
                        empresa=compra.empresa,
                        data=compra.data_emissao,
                        descricao=f"Compra {compra.numero} - Registo",
                        compra=compra
                    )
                    
                    # 1. Débito em Gastos (Base)
                    # Se tiver imposto dedutível? Simplificado: Lança Total em Gastos se não tiver IVA separado.
                    # Se tiver retenção, o gasto é o valor TOTAL do serviço, não líquido.
                    
                    MovimentoRazao.objects.create(lancamento=lancamento, conta=conta_gastos, tipo='D', valor=compra.subtotal)
                    
                    if compra.total_imposto > 0:
                         # Tenta achar IVA Dedutivel
                        conta_iva = Conta.objects.filter(codigo='34.5.2', empresa=compra.empresa).first() 
                        if not conta_iva: conta_iva = Conta.objects.filter(codigo__startswith='34', empresa=compra.empresa).first()
                        MovimentoRazao.objects.create(lancamento=lancamento, conta=conta_iva, tipo='D', valor=compra.total_imposto)
                    
                    # 2. Crédito no Fornecedor (Paga apenas o líquido: Total - Retenção)
                    MovimentoRazao.objects.create(lancamento=lancamento, conta=conta_fornecedor, tipo='C', valor=compra.total_a_pagar)
                    
                    # 3. Crédito na Retenção na Fonte a Pagar (Estado)
                    if compra.valor_retencao > 0:
                        conta_retencao_pagar = Conta.objects.filter(codigo='34.1', empresa=compra.empresa).first() # Estado - Impostos a Pagar
                        if not conta_retencao_pagar: conta_retencao_pagar = Conta.objects.filter(codigo__startswith='34', empresa=compra.empresa).first()
                        
                        if conta_retencao_pagar:
                            MovimentoRazao.objects.create(lancamento=lancamento, conta=conta_retencao_pagar, tipo='C', valor=compra.valor_retencao)

            except Exception as e:
                pass 
                
        messages.success(request, f'Estado da compra {compra.numero} alterado para {estado}.')
    return redirect('compras_list')

from .utils_pdf import gerar_pdf_fatura

def pdf_fatura(request, fatura_id):
    return gerar_pdf_fatura(request, fatura_id)

def despesas_list(request):
    """Lista de Despesas e Formulário de Adição"""
    empresa = request.empresa
    despesas = Despesa.objects.filter(empresa=empresa).order_by('-id')
    
    if request.method == 'POST':
        form = DespesaForm(request.POST, empresa=empresa)
        if form.is_valid():
            despesa = form.save(commit=False)
            despesa.empresa = empresa
            despesa.save()
            messages.success(request, f'Despesa {despesa.numero} registada com sucesso!')
            return redirect('despesas')
        else:
            messages.error(request, 'Erro ao registar despesa. Verifique os campos.')
    else:
        form = DespesaForm(empresa=empresa)

    return render(request, 'despesas.html', {
        'despesas': despesas,
        'form': form
    })

@user_passes_test(lambda u: u.is_superuser)
def usuarios_list(request):
    """Lista de Usuários no módulo de Administração"""
    usuarios = User.objects.all()
    return render(request, 'admin/usuarios_list.html', {'usuarios': usuarios})

@user_passes_test(lambda u: u.is_superuser)
def novo_usuario(request):
    """Criação de novos usuários com permissões"""
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuário criado com sucesso.')
            return redirect('usuarios_list')
    else:
        form = UsuarioForm()
    return render(request, 'admin/usuario_form.html', {'form': form, 'title': 'Novo Usuário'})

@user_passes_test(lambda u: u.is_superuser)
def editar_usuario(request, pk):
    """Edição de utilizadores existentes"""
    usuario = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Utilizador atualizado com sucesso.')
            return redirect('usuarios_list')
    else:
        form = UsuarioForm(instance=usuario)
    return render(request, 'admin/usuario_form.html', {'form': form, 'title': f'Editar: {usuario.username}'})

@user_passes_test(lambda u: u.is_superuser)
def excluir_usuario(request, pk):
    """Exclui um usuário"""
    if request.user.pk == pk:
        messages.error(request, 'Você não pode excluir o seu próprio usuário.')
        return redirect('usuarios_list')
    usuario = get_object_or_404(User, pk=pk)
    usuario.delete()
    messages.success(request, 'Usuário excluído com sucesso.')
    return redirect('usuarios_list')

def configuracao_empresa(request):
    """Gestão de empresas (Multi-empresa)"""
    empresas = Empresa.objects.all()
    
    if request.method == 'POST':
        form = EmpresaForm(request.POST, request.FILES)
        if form.is_valid():
            nova_empresa = form.save()
            messages.success(request, f'Empresa {nova_empresa.nome} criada com sucesso!')
            
            # Inicializar plano de contas automaticamente se selecionado
            if nova_empresa.plano_modelo != 'PERSONALIZADO':
                count = _executar_inicializacao_plano(nova_empresa)
                if count > 0:
                    messages.info(request, f'Plano de contas {nova_empresa.get_plano_modelo_display()} configurado ({count} contas).')

            # Se for a única empresa, selecionar automaticamente
            if Empresa.objects.count() == 1:
                request.session['empresa_id'] = nova_empresa.id
            return redirect('configuracao_empresa')
    else:
        form = EmpresaForm()
        
    return render(request, 'empresa_config.html', {
        'form': form,
        'empresas': empresas,
        'active_empresa': request.empresa,
        'bna_rates': _get_bna_rates()
    })

def selecionar_empresa(request, pk):
    """Muda a empresa ativa na sessão"""
    try:
        empresa = Empresa.objects.get(pk=pk)
        request.session['empresa_id'] = empresa.id
        messages.success(request, f'Trabalhando agora em: {empresa.nome}')
    except Empresa.DoesNotExist:
        messages.error(request, 'Empresa não encontrada.')
    
    return redirect(request.META.get('HTTP_REFERER', 'home'))

def editar_empresa(request, pk):
    """Edita dados de uma empresa específica"""
    empresa = Empresa.objects.get(pk=pk)
    if request.method == 'POST':
        form = EmpresaForm(request.POST, request.FILES, instance=empresa)
        if form.is_valid():
            form.save()
            messages.success(request, f'Dados de {empresa.nome} atualizados.')
            return redirect('configuracao_empresa')
    else:
        form = EmpresaForm(instance=empresa)
    
    return render(request, 'empresa_form.html', {
        'form': form, 
        'empresa': empresa,
        'bna_rates': _get_bna_rates()
    })

def limpar_dados_empresa(request):
    """
    Remove todos os dados transacionais e registos (Faturas, Compras, Lançamentos, Stocks, RH)
    mas mantém a Empresa e o seu Plano de Contas.
    """
    empresa = request.empresa
    if not empresa:
        messages.error(request, "Selecione uma empresa primeiro.")
        return redirect('configuracao_empresa')

    if request.method == 'POST':
        from .models import (
            Fatura, Compra, Despesa, LancamentoDiario, 
            ProcessamentoSalarial, Falta, HoraExtra,
            MovimentoStock, Inventario, Cliente, Fornecedor, Produto, Funcionario
        )
        
        # Ordem de eliminação para respeitar FKs
        # 1. Transações de RH
        Falta.objects.filter(funcionario__empresa=empresa).delete()
        HoraExtra.objects.filter(funcionario__empresa=empresa).delete()
        ProcessamentoSalarial.objects.filter(funcionario__empresa=empresa).delete()
        
        # 2. Transações de Stocks
        MovimentoStock.objects.filter(empresa=empresa).delete()
        Inventario.objects.filter(empresa=empresa).delete()
        
        # 3. Transações Financeiras
        # Primeiro desassociar FKs de lançamentos
        Fatura.objects.filter(empresa=empresa).update(lancamento=None)
        Compra.objects.filter(empresa=empresa).update(lancamento=None)
        Despesa.objects.filter(empresa=empresa).update(lancamento=None)
        ProcessamentoSalarial.objects.filter(funcionario__empresa=empresa).update(lancamento=None)
        
        LancamentoDiario.objects.filter(empresa=empresa).delete()
        Fatura.objects.filter(empresa=empresa).delete()
        Compra.objects.filter(empresa=empresa).delete()
        Despesa.objects.filter(empresa=empresa).delete()
        
        # 4. Entidades e Cadastros (Opcional, mas geralmente usuário quer resetar tudo)
        Cliente.objects.filter(empresa=empresa).delete()
        Fornecedor.objects.filter(empresa=empresa).delete()
        Funcionario.objects.filter(empresa=empresa).delete()
        Produto.objects.filter(empresa=empresa).delete()
        
        messages.success(request, f"Todos os dados de '{empresa.nome}' foram limpos com sucesso. O Plano de Contas foi preservado.")
        return redirect('configuracao_empresa')
        
    return redirect('configuracao_empresa')


def lista_cambios(request):
    """Lista e adiciona novos câmbios ao histórico"""
    from .forms import CambioForm
    from .models import Cambio
    empresa = request.empresa
    if not empresa:
        messages.error(request, "Selecione uma empresa primeiro.")
        return redirect('configuracao_empresa')
        
    historico = Cambio.objects.filter(empresa=empresa).order_by('-data_inicio')
    
    if request.method == 'POST':
        form = CambioForm(request.POST)
        if form.is_valid():
            novo_cambio = form.save(commit=False)
            novo_cambio.empresa = empresa
            novo_cambio.save()
            messages.success(request, f'Nova taxa de câmbio adicionada para {novo_cambio.data_inicio}.')
            return redirect('lista_cambios')
    else:
        form = CambioForm()
        
    return render(request, 'cambios_lista.html', {
        'form': form,
        'historico': historico,
    })

def eliminar_cambio(request, pk):
    """Remove uma taxa do histórico"""
    from .models import Cambio
    cambio = get_object_or_404(Cambio, pk=pk, empresa=request.empresa)
    cambio.delete()
    messages.success(request, "Taxa de câmbio removida.")
    return redirect('lista_cambios')

def lancamento_abertura(request):
    """View para o lançamento de abertura de exercício"""
    if request.method == 'POST':
        form = LancamentoAberturaForm(request.POST)
        formset = MovimentoAberturaFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            abertura = form.save(commit=False)
            abertura.tipo = 'ABERTURA'
            abertura.save()
            
            movimentos = formset.save(commit=False)
            total_debito = 0
            total_credito = 0
            
            for mov in movimentos:
                mov.lancamento = abertura
                mov.save()
                if mov.tipo == 'D': total_debito += mov.valor
                else: total_credito += mov.valor
                
            # Validar partida dobrada (Débito = Crédito)
            if total_debito != total_credito:
                messages.warning(request, f'Atenção: O balanço de abertura não está equilibrado (D:{total_debito} / C:{total_credito}).')
            
            messages.success(request, 'Lançamento de abertura registado com sucesso!')
            return redirect('relatorio_diario')
    else:
        # Sugerir data de início do ano da empresa
        empresa = Empresa.objects.first()
        ano = empresa.ano_exercicio if empresa else 2025
        data_sugerida = f"{ano}-01-01"
        
        form = LancamentoAberturaForm(initial={'data': data_sugerida, 'descricao': f'Saldos de Abertura - {ano}'})
        formset = MovimentoAberturaFormSet()

    return render(request, 'lancamento_abertura.html', {
        'form': form,
        'formset': formset
    })

def lancamentos_list(request):
    """Lista de todos os lançamentos manuais"""
    empresa = request.empresa
    lancamentos = LancamentoDiario.objects.filter(
        empresa=empresa,
        fatura__isnull=True, 
        compra__isnull=True, 
        despesa__isnull=True
    ).order_by('-data', '-id')
    return render(request, 'lancamentos_lista.html', {'lancamentos': lancamentos})

def novo_lancamento(request):
    """Cria um novo lançamento manual"""
    from .forms import LancamentoManualForm, MovimentoManualFormSet
    empresa = request.empresa
    
    if request.method == 'POST':
        form = LancamentoManualForm(request.POST)
        formset = MovimentoManualFormSet(request.POST, form_kwargs={'empresa': empresa})
        
        if form.is_valid() and formset.is_valid():
            lancamento = form.save(commit=False)
            lancamento.empresa = empresa
            lancamento.save()
            
            movimentos = formset.save(commit=False)
            
            total_debito = 0
            total_credito = 0
            
            for mov in movimentos:
                mov.lancamento = lancamento
                mov.save()
                if mov.tipo == 'D': total_debito += mov.valor
                else: total_credito += mov.valor
                
            if total_debito != total_credito:
                messages.warning(request, f'Lançamento gravado, mas está desequilibrado! (D: {total_debito} / C: {total_credito})')
            else:
                messages.success(request, 'Lançamento efetuado com sucesso!')
                
            return redirect('lancamentos_list')
    else:
        form = LancamentoManualForm()
        formset = MovimentoManualFormSet(form_kwargs={'empresa': empresa})
        
    return render(request, 'lancamento_manual_form.html', {
        'form': form,
        'formset': formset
    })

def lancamento_detalhe(request, pk):
    """Visualiza detalhes de um lançamento específico"""
    from .models import LancamentoDiario
    lancamento = LancamentoDiario.objects.prefetch_related('movimentos__conta').get(pk=pk)
    return render(request, 'lancamento_detalhe.html', {'lancamento': lancamento})

def plano_contas(request):
    """Lista o Plano Geral de Contabilidade"""
    empresa = request.empresa
    if empresa:
        # Se tem empresa, mostra apenas o plano da empresa
        contas = Conta.objects.filter(empresa=empresa).order_by('codigo')
    else:
        # Se não tem empresa (admin), mostra o plano global
        contas = Conta.objects.filter(empresa__isnull=True).order_by('codigo')
        
    return render(request, 'plano_contas.html', {'contas': contas})

def nova_conta(request):
    """Cria uma nova conta no PGC"""
    from .forms import ContaForm
    if request.method == 'POST':
        form = ContaForm(request.POST, empresa=request.empresa)
        if form.is_valid():
            conta = form.save(commit=False)
            conta.empresa = request.empresa
            conta.save()
            messages.success(request, f'Conta {conta.codigo} criada com sucesso!')
            return redirect('plano_contas')
    else:
        form = ContaForm(empresa=request.empresa)
    
    return render(request, 'conta_form.html', {'form': form, 'titulo': 'Nova Conta'})

def editar_conta(request, pk):
    """Edita uma conta existente"""
    from .forms import ContaForm
    conta = get_object_or_404(Conta, pk=pk)
    
    if request.method == 'POST':
        form = ContaForm(request.POST, instance=conta, empresa=request.empresa)
        if form.is_valid():
            form.save()
            messages.success(request, f'Conta {conta.codigo} atualizada com sucesso!')
            return redirect('plano_contas')
        else:
            messages.error(request, 'Não foi possível gravar as alterações. Verifique os erros no formulário.')
    else:
        form = ContaForm(instance=conta, empresa=request.empresa)
    
    return render(request, 'conta_form.html', {'form': form, 'titulo': f'Editar Conta: {conta.codigo}', 'conta': conta})

def eliminar_conta(request, pk):
    """Exclui uma conta e as suas subcontas se não houver movimentos"""
    conta = get_object_or_404(Conta, pk=pk)
    empresa = request.empresa
    
    # Segurança: Não permitir eliminar contas globais a partir de uma sessão de empresa
    if empresa and conta.empresa is None:
        messages.error(request, 'Não tem permissão para eliminar contas do Plano mestre do sistema.')
        return redirect('plano_contas')

    # Função interna para verificar movimentos recursivamente
    def tem_movimentos_recursivo(c):
        if c.movimentos.exists():
            return True
        for sub in c.subcontas.all():
            if tem_movimentos_recursivo(sub):
                return True
        return False

    if tem_movimentos_recursivo(conta):
        messages.error(request, f'Impossível eliminar: a conta {conta.codigo} (ou as suas subcontas) já possuem movimentos registados no diário.')
        return redirect('plano_contas')
    
    # Se chegamos aqui, podemos eliminar a conta e toda a sua árvore de subcontas
    contas_para_eliminar = []
    def coletar_contas(c):
        contas_para_eliminar.append(c)
        for sub in c.subcontas.all():
            coletar_contas(sub)
    
    coletar_contas(conta)
    total = len(contas_para_eliminar)
    
    for c in reversed(contas_para_eliminar): # Elimina de baixo para cima
        c.delete()
        
    messages.success(request, f'Conta {conta.codigo} e as suas {total-1} subcontas foram eliminadas com sucesso.')
    return redirect('plano_contas')

def pdf_plano_contas(request):
    """Gera o PDF do Plano de Contas"""
    from .utils_pdf import gerar_pdf_plano_contas
    return gerar_pdf_plano_contas(request)

def importar_plano_contas(request):
    """Importa o Plano de Contas a partir de TXT ou Excel"""
    from .forms import ImportarPGCForm
    from .models import Conta, Classe
    import re
    import openpyxl
    
    if request.method == 'POST':
        form = ImportarPGCForm(request.POST, request.FILES)
        if form.is_valid():
            arquivo = request.FILES['arquivo']
            filename = arquivo.name.lower()
            empresa = request.empresa
            
            count_created = 0
            try:
                if filename.endswith('.txt'):
                    content = arquivo.read().decode('utf-8')
                    lines = content.splitlines()
                    for line in lines:
                        line = line.strip()
                        if not line: continue
                        
                        match = re.search(r'^(.*?)\s+(.*?)\s+([RIMA])$', line)
                        if match:
                            codigo = match.group(1).strip()
                            descricao = match.group(2).strip()
                            tipo_char = match.group(3).strip()
                            
                            if process_conta_import(codigo, descricao, tipo_char, empresa):
                                count_created += 1
                
                elif filename.endswith(('.xlsx', '.xlsm')):
                    wb = openpyxl.load_workbook(arquivo, data_only=True)
                    sheet = wb.active
                    for row in sheet.iter_rows(min_row=1, values_only=True):
                        # Espera formato: Codigo, Descricao, Tipo (R/I/M)
                        if len(row) >= 3 and row[0] and row[1]:
                            codigo = str(row[0]).strip()
                            descricao = str(row[1]).strip()
                            tipo_char = str(row[2]).strip().upper() if row[2] else 'M'
                            
                            if process_conta_import(codigo, descricao, tipo_char, empresa):
                                count_created += 1
                
                elif filename.endswith('.xls'):
                     raise Exception("Formato .xls (antigo) não suportado. Por favor guarde como .xlsx e tente novamente.")
                
                else:
                     raise Exception("Formato de ficheiro desconhecido. Use .xlsx (Excel) ou .txt")
                
                messages.success(request, f'Importação concluída! {count_created} contas processadas.')
                return redirect('plano_contas')
                
            except Exception as e:
                messages.error(request, f'Erro durante a importação: {str(e)}')
    else:
        form = ImportarPGCForm()
        
    return render(request, 'importar_pgc.html', {'form': form})

def process_conta_import(codigo, descricao, tipo_char, empresa):
    """Helper para processar uma única conta na importação"""
    from .models import Conta, Classe
    tipo_map = {
        'R': 'RAZAO', 'RAZÃO': 'RAZAO', 'RAZAO': 'RAZAO',
        'I': 'INTEGRACAO', 'INTEGRAÇÃO': 'INTEGRACAO', 'INTEGRACAO': 'INTEGRACAO',
        'M': 'MOVIMENTO', 'MOVIMENTO': 'MOVIMENTO',
        'A': 'APURAMENTO', 'APURAMENTO': 'APURAMENTO'
    }
    # Sanitize: Remove espaços, underscores e garantir upper
    clean_char = tipo_char.upper().replace('_', ' ').strip()
    # Tenta match direto ou pega primeira letra se não encontrar (fallback inteligente)
    if clean_char in tipo_map:
        tipo = tipo_map[clean_char]
    elif clean_char and clean_char[0] in tipo_map:
         tipo = tipo_map[clean_char[0]]
    else:
        tipo = 'MOVIMENTO'
    
    try:
        classe_cod = codigo[0]
        classe = Classe.objects.get(codigo=classe_cod)
        
        # Encontrar pai
        pai = None
        potential_parent_code = codigo[:-1]
        while len(potential_parent_code) >= 2:
            pai = Conta.objects.filter(codigo=potential_parent_code, empresa=empresa).first()
            if not pai: # Tentar global se não for da empresa
                pai = Conta.objects.filter(codigo=potential_parent_code, empresa__isnull=True).first()
            if pai: break
            potential_parent_code = potential_parent_code[:-1]

        Conta.objects.update_or_create(
            codigo=codigo,
            empresa=empresa,
            defaults={
                'descricao': descricao,
                'tipo': tipo,
                'aceita_lancamentos': (tipo in ['MOVIMENTO', 'APURAMENTO']),
                'classe': classe,
                'conta_pai': pai
            }
        )
        return True
    except Exception:
        return False

def inicializar_plano(request):
    """Inicializa o plano de contas da empresa com base no modelo selecionado"""
    empresa = request.empresa
    if not empresa:
        messages.error(request, "Selecione uma empresa primeiro.")
        return redirect('configuracao_empresa')
        
    if Conta.objects.filter(empresa=empresa).exists():
        messages.warning(request, "A empresa já possui contas registadas. Operação cancelada.")
        return redirect('plano_contas')
        
    contas_criadas = _executar_inicializacao_plano(empresa)
    
    if contas_criadas > 0:
        messages.success(request, f"Plano inicializado: {contas_criadas} contas criadas.")
    else:
        messages.error(request, "O Plano Geral global não está carregado ou ocorreu um erro.")
        
    return redirect('plano_contas')

def _executar_inicializacao_plano(empresa):
    """Lógica principal de cópia do plano global para uma empresa"""
    contas_base = Conta.objects.filter(empresa__isnull=True).order_by('codigo')
    if not contas_base.exists():
        return 0
        
    contas_criadas = 0
    for c_base in contas_base:
        pai_novo = None
        if c_base.conta_pai:
            pai_novo = Conta.objects.filter(codigo=c_base.conta_pai.codigo, empresa=empresa).first()
            
        Conta.objects.create(
            codigo=c_base.codigo,
            descricao=c_base.descricao,
            tipo=c_base.tipo,
            tipo_entidade=c_base.tipo_entidade,
            aceita_lancamentos=c_base.aceita_lancamentos,
            classe=c_base.classe,
            conta_pai=pai_novo,
            empresa=empresa
        )
        contas_criadas += 1
    return contas_criadas

def exportar_plano_contas_excel(request):
    """Exporta o Plano de Contas para Excel"""
    import openpyxl
    from django.http import HttpResponse
    from django.utils import timezone
    from openpyxl.styles import Font, Alignment
    
    empresa = request.empresa
    if empresa:
        # Se tem empresa, mostra apenas o plano da empresa
        contas = Conta.objects.filter(empresa=empresa).order_by('codigo')
        filename = f"Plano_Contas_{empresa.nome}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    else:
        # Se não tem empresa (admin), mostra o plano global
        contas = Conta.objects.filter(empresa__isnull=True).order_by('codigo')
        filename = f"Plano_Contas_Global_{timezone.now().strftime('%Y%m%d')}.xlsx"
        
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plano de Contas"
    
    # Cabeçalho
    headers = ["Código", "Descrição", "Tipo", "Classe", "Natureza"]
    ws.append(headers)
    
    # Estilização do cabeçalho
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
    # Dados
    for conta in contas:
        ws.append([
            conta.codigo,
            conta.descricao,
            conta.get_tipo_display(),
            str(conta.classe),
            conta.get_tipo_entidade_display() if conta.tipo_entidade != 'NENHUM' else ''
        ])
        
    # Ajuste de largura das colunas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 20
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

def contas_correntes_list(request):
    """Lista de Contas Correntes (Clientes e Fornecedores)"""
    from .relatorios import calcular_saldo_conta
    from decimal import Decimal
    empresa = request.empresa
    if not empresa: return redirect('configuracao_empresa')

    # 1. Clientes
    clientes = Cliente.objects.filter(empresa=empresa)
    lista_clientes = []
    for c in clientes:
        saldo = Decimal(0)
        conta = Conta.objects.filter(codigo=c.codigo_contabilistico, empresa=empresa).first()
        if conta:
            dados = calcular_saldo_conta(conta, empresa)
            saldo = dados['saldo'] 
        lista_clientes.append({'id': c.id, 'nome': c.nome, 'nif': c.nif, 'codigo_contabilistico': c.codigo_contabilistico, 'saldo': saldo})

    # 2. Fornecedores
    fornecedores = Fornecedor.objects.filter(empresa=empresa)
    lista_fornecedores = []
    for f in fornecedores:
        saldo = Decimal(0)
        conta = Conta.objects.filter(codigo=f.codigo_contabilistico, empresa=empresa).first()
        if conta:
            dados = calcular_saldo_conta(conta, empresa)
            # Para fornecedores, queremos mostrar o valor que devemos (Credor)
            saldo = -dados['saldo'] 
        lista_fornecedores.append({'id': f.id, 'nome': f.nome, 'nif': f.nif, 'codigo_contabilistico': f.codigo_contabilistico, 'saldo': saldo})

    return render(request, 'contas_correntes_list.html', {
        'lista_clientes': lista_clientes,
        'lista_fornecedores': lista_fornecedores
    })

def conta_corrente_detalhe(request, tipo, pk):
    """Extrato detalhado"""
    from .relatorios import gerar_razao
    empresa = request.empresa
    
    if tipo == 'cliente':
        entidade = get_object_or_404(Cliente, pk=pk)
    elif tipo == 'fornecedor':
        entidade = get_object_or_404(Fornecedor, pk=pk)
    else:
        return redirect('contas_correntes_list')
        
    conta = Conta.objects.filter(codigo=entidade.codigo_contabilistico, empresa=empresa).first()
    if not conta:
        messages.error(request, f'Conta não encontrada: {entidade.codigo_contabilistico}')
        return redirect('contas_correntes_list')
        
    dados = gerar_razao(conta.id, empresa)
    
    if tipo == 'fornecedor':
        dados['saldo_anterior'] = -dados['saldo_anterior']
        dados['saldo_final'] = -dados['saldo_final']
        for linha in dados['linhas']:
            linha['saldo'] = -linha['saldo']
            
    return render(request, 'conta_corrente_detalhe.html', {
        'entidade': entidade,
        'movimentos': dados['linhas'],
        'saldo_anterior': dados['saldo_anterior'],
        'saldo_final': dados['saldo_final'],
        'total_debito': dados['total_debito'],
        'total_credito': dados['total_credito'],
    })
# ===== EXPORTAÇÃO DE CLIENTES E FORNECEDORES =====

def exportar_clientes_excel(request):
    """Exportar lista de clientes para Excel"""
    empresa = request.empresa
    clientes = Cliente.objects.filter(empresa=empresa).order_by('codigo_contabilistico')
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return exportar_clientes_csv(request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Headers
    headers = ["Código PGC", "Nome", "NIF", "Email", "Telefone", "Endereço"]
    ws.append(headers)
    
    # Estilo do header
    header_fill = PatternFill(start_color="58A6FF", end_color="58A6FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dados
    for cliente in clientes:
        ws.append([
            cliente.codigo_contabilistico,
            cliente.nome,
            cliente.nif or "",
            cliente.email or "",
            cliente.telefone or "",
            cliente.endereco or "",
        ])
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    
    # Retornar arquivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response

def exportar_clientes_csv(request):
    """Exportar lista de clientes para CSV (fallback)"""
    import csv
    empresa = request.empresa
    clientes = Cliente.objects.filter(empresa=empresa).order_by('codigo_contabilistico')
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="Clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(["Código PGC", "Nome", "NIF", "Email", "Telefone", "Endereço"])
    
    for cliente in clientes:
        writer.writerow([
            cliente.codigo_contabilistico,
            cliente.nome,
            cliente.nif or "",
            cliente.email or "",
            cliente.telefone or "",
            cliente.endereco or "",
        ])
    
    return response

def exportar_clientes_pdf(request):
    """Exportar lista de clientes para PDF"""
    empresa = request.empresa
    clientes = Cliente.objects.filter(empresa=empresa).order_by('codigo_contabilistico')
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
    except ImportError:
        return HttpResponse("ReportLab não está instalado. Instale com: pip install reportlab", status=400)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#58A6FF'),
        spaceAfter=12,
        alignment=1  # Center
    )
    elements.append(Paragraph(f"Lista de Clientes - {empresa.nome}", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Tabela
    data = [["Código PGC", "Nome", "NIF", "Telefone"]]
    for cliente in clientes:
        data.append([
            cliente.codigo_contabilistico,
            cliente.nome[:30],  # Limitar comprimento
            cliente.nif or "",
            cliente.telefone or "",
        ])
    
    table = Table(data, colWidths=[1.2*inch, 2*inch, 1.2*inch, 1.3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#58A6FF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response

def exportar_fornecedores_excel(request):
    """Exportar lista de fornecedores para Excel"""
    empresa = request.empresa
    fornecedores = Fornecedor.objects.filter(empresa=empresa).order_by('codigo_contabilistico')
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return exportar_fornecedores_csv(request)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fornecedores"
    
    # Headers
    headers = ["Código PGC", "Nome", "NIF", "Email", "Telefone", "Endereço"]
    ws.append(headers)
    
    # Estilo do header
    header_fill = PatternFill(start_color="34A853", end_color="34A853", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dados
    for fornecedor in fornecedores:
        ws.append([
            fornecedor.codigo_contabilistico,
            fornecedor.nome,
            fornecedor.nif or "",
            fornecedor.email or "",
            fornecedor.telefone or "",
            fornecedor.endereco or "",
        ])
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    
    # Retornar arquivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Fornecedores_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response

def exportar_fornecedores_csv(request):
    """Exportar lista de fornecedores para CSV (fallback)"""
    import csv
    empresa = request.empresa
    fornecedores = Fornecedor.objects.filter(empresa=empresa).order_by('codigo_contabilistico')
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="Fornecedores_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(["Código PGC", "Nome", "NIF", "Email", "Telefone", "Endereço"])
    
    for fornecedor in fornecedores:
        writer.writerow([
            fornecedor.codigo_contabilistico,
            fornecedor.nome,
            fornecedor.nif or "",
            fornecedor.email or "",
            fornecedor.telefone or "",
            fornecedor.endereco or "",
        ])
    
    return response

def exportar_fornecedores_pdf(request):
    """Exportar lista de fornecedores para PDF"""
    empresa = request.empresa
    fornecedores = Fornecedor.objects.filter(empresa=empresa).order_by('codigo_contabilistico')
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
    except ImportError:
        return HttpResponse("ReportLab não está instalado. Instale com: pip install reportlab", status=400)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#34A853'),
        spaceAfter=12,
        alignment=1  # Center
    )
    elements.append(Paragraph(f"Lista de Fornecedores - {empresa.nome}", title_style))
    elements.append(Spacer(1, 0.3*inch))
    
    # Tabela
    data = [["Código PGC", "Nome", "NIF", "Telefone"]]
    for fornecedor in fornecedores:
        data.append([
            fornecedor.codigo_contabilistico,
            fornecedor.nome[:30],  # Limitar comprimento
            fornecedor.nif or "",
            fornecedor.telefone or "",
        ])
    
    table = Table(data, colWidths=[1.2*inch, 2*inch, 1.2*inch, 1.3*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34A853')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Fornecedores_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response


# ===== EXPORTAÇÃO DE CONTAS CORRENTES =====

def exportar_conta_corrente_excel(request, pk):
    """Exportar extrato da conta corrente para Excel"""
    from .relatorios import gerar_razao
    empresa = request.empresa
    
    # Determinar se é cliente ou fornecedor
    cliente = Cliente.objects.filter(pk=pk, empresa=empresa).first()
    fornecedor = Fornecedor.objects.filter(pk=pk, empresa=empresa).first()
    
    if cliente:
        entidade = cliente
        tipo = 'Cliente'
    elif fornecedor:
        entidade = fornecedor
        tipo = 'Fornecedor'
    else:
        return HttpResponse('Entidade não encontrada', status=404)
    
    conta = Conta.objects.filter(codigo=entidade.codigo_contabilistico, empresa=empresa).first()
    if not conta:
        return HttpResponse('Conta não encontrada', status=404)
    
    dados = gerar_razao(conta.id, empresa)
    
    if isinstance(entidade, Fornecedor):
        dados['saldo_anterior'] = -dados['saldo_anterior']
        dados['saldo_final'] = -dados['saldo_final']
        for linha in dados['linhas']:
            linha['saldo'] = -linha['saldo']
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return exportar_conta_corrente_csv(request, pk)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Extrato"
    
    # Cabeçalho
    ws['A1'] = f"Extrato de Conta Corrente - {tipo}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"{entidade.nome}"
    ws['A2'].font = Font(bold=True, size=11)
    ws.merge_cells('A2:F2')
    
    ws['A3'] = f"NIF: {entidade.nif} | Código: {entidade.codigo_contabilistico}"
    ws.merge_cells('A3:F3')
    
    # Headers da tabela
    headers = ["Data", "Documento", "Descrição", "Débito", "Crédito", "Saldo"]
    ws.append([])  # Linha em branco
    ws.append(headers)
    
    # Estilo do header
    header_fill = PatternFill(start_color="58A6FF", end_color="58A6FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col in range(1, 7):
        cell = ws.cell(row=5, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Saldo Anterior
    row = 6
    ws[f'A{row}'] = "Saldo Anterior"
    ws[f'F{row}'] = dados['saldo_anterior']
    ws[f'F{row}'].number_format = '#,##0.00'
    
    # Dados
    row = 7
    for linha in dados['linhas']:
        ws[f'A{row}'] = linha['data'].strftime('%d/%m/%Y')
        ws[f'B{row}'] = linha.get('documento', '')
        ws[f'C{row}'] = linha['descricao']
        ws[f'D{row}'] = linha['debito'] if linha['debito'] > 0 else ''
        ws[f'E{row}'] = linha['credito'] if linha['credito'] > 0 else ''
        ws[f'F{row}'] = linha['saldo']
        
        ws[f'D{row}'].number_format = '#,##0.00'
        ws[f'E{row}'].number_format = '#,##0.00'
        ws[f'F{row}'].number_format = '#,##0.00'
        
        row += 1
    
    # Totais
    row += 1
    ws[f'A{row}'] = "TOTAIS"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'D{row}'] = dados['total_debito']
    ws[f'E{row}'] = dados['total_credito']
    ws[f'D{row}'].font = Font(bold=True)
    ws[f'E{row}'].font = Font(bold=True)
    ws[f'D{row}'].number_format = '#,##0.00'
    ws[f'E{row}'].number_format = '#,##0.00'
    
    # Saldo Final
    row += 1
    ws[f'A{row}'] = "Saldo Final"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'F{row}'] = dados['saldo_final']
    ws[f'F{row}'].font = Font(bold=True)
    ws[f'F{row}'].number_format = '#,##0.00'
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Extrato_{entidade.nome.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response

def exportar_conta_corrente_csv(request, pk):
    """Exportar extrato para CSV (fallback)"""
    import csv
    from .relatorios import gerar_razao
    empresa = request.empresa
    
    cliente = Cliente.objects.filter(pk=pk, empresa=empresa).first()
    fornecedor = Fornecedor.objects.filter(pk=pk, empresa=empresa).first()
    
    if cliente:
        entidade = cliente
    elif fornecedor:
        entidade = fornecedor
    else:
        return HttpResponse('Entidade não encontrada', status=404)
    
    conta = Conta.objects.filter(codigo=entidade.codigo_contabilistico, empresa=empresa).first()
    if not conta:
        return HttpResponse('Conta não encontrada', status=404)
    
    dados = gerar_razao(conta.id, empresa)
    
    if isinstance(entidade, Fornecedor):
        dados['saldo_anterior'] = -dados['saldo_anterior']
        dados['saldo_final'] = -dados['saldo_final']
        for linha in dados['linhas']:
            linha['saldo'] = -linha['saldo']
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="Extrato_{entidade.nome.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([f"Extrato de Conta Corrente - {entidade.nome}"])
    writer.writerow([f"NIF: {entidade.nif} | Código: {entidade.codigo_contabilistico}"])
    writer.writerow([])
    writer.writerow(["Data", "Documento", "Descrição", "Débito", "Crédito", "Saldo"])
    writer.writerow([f"Saldo Anterior", "", "", "", "", f"{dados['saldo_anterior']:.2f}"])
    
    for linha in dados['linhas']:
        writer.writerow([
            linha['data'].strftime('%d/%m/%Y'),
            linha.get('documento', ''),
            linha['descricao'],
            f"{linha['debito']:.2f}" if linha['debito'] > 0 else "",
            f"{linha['credito']:.2f}" if linha['credito'] > 0 else "",
            f"{linha['saldo']:.2f}"
        ])
    
    writer.writerow([])
    writer.writerow(["TOTAIS", "", "", f"{dados['total_debito']:.2f}", f"{dados['total_credito']:.2f}", ""])
    writer.writerow(["Saldo Final", "", "", "", "", f"{dados['saldo_final']:.2f}"])
    
    return response

def exportar_conta_corrente_pdf(request, pk):
    """Exportar extrato para PDF"""
    from .relatorios import gerar_razao
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    
    empresa = request.empresa
    
    cliente = Cliente.objects.filter(pk=pk, empresa=empresa).first()
    fornecedor = Fornecedor.objects.filter(pk=pk, empresa=empresa).first()
    
    if cliente:
        entidade = cliente
        cor = '#58A6FF'
    elif fornecedor:
        entidade = fornecedor
        cor = '#34A853'
    else:
        return HttpResponse('Entidade não encontrada', status=404)
    
    conta = Conta.objects.filter(codigo=entidade.codigo_contabilistico, empresa=empresa).first()
    if not conta:
        return HttpResponse('Conta não encontrada', status=404)
    
    dados = gerar_razao(conta.id, empresa)
    
    if isinstance(entidade, Fornecedor):
        dados['saldo_anterior'] = -dados['saldo_anterior']
        dados['saldo_final'] = -dados['saldo_final']
        for linha in dados['linhas']:
            linha['saldo'] = -linha['saldo']
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        textColor=colors.HexColor(cor),
        spaceAfter=6,
        alignment=1
    )
    elements.append(Paragraph(f"Extrato de Conta Corrente", title_style))
    
    # Dados da entidade
    info_style = ParagraphStyle(
        'Info',
        parent=styles['Normal'],
        fontSize=10,
        alignment=1,
        spaceAfter=12
    )
    elements.append(Paragraph(f"<b>{entidade.nome}</b><br/>NIF: {entidade.nif} | Código: {entidade.codigo_contabilistico}", info_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Tabela
    data = [["Data", "Documento", "Descrição", "Débito", "Crédito", "Saldo"]]
    data.append(["Saldo Anterior", "", "", "", "", f"{dados['saldo_anterior']:.2f}"])
    
    for linha in dados['linhas']:
        data.append([
            linha['data'].strftime('%d/%m/%Y'),
            linha.get('documento', ''),
            linha['descricao'][:25],
            f"{linha['debito']:.2f}" if linha['debito'] > 0 else "",
            f"{linha['credito']:.2f}" if linha['credito'] > 0 else "",
            f"{linha['saldo']:.2f}"
        ])
    
    data.append(["TOTAIS", "", "", f"{dados['total_debito']:.2f}", f"{dados['total_credito']:.2f}", ""])
    data.append(["Saldo Final", "", "", "", "", f"{dados['saldo_final']:.2f}"])
    
    table = Table(data, colWidths=[0.9*inch, 1*inch, 1.8*inch, 0.9*inch, 0.9*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(cor)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BACKGROUND', (0, 1), (-1, 1), colors.lightgrey),
        ('BACKGROUND', (0, -2), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Extrato_{entidade.nome.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response